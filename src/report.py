"""
Generator benchmark reportov pre konkretnu firmu.

Vytvara Excel (.xlsx) a HTML (.html) report.

Pouzitie:
    python -m src.report --ico 31321828
"""

import argparse
from pathlib import Path

import pandas as pd
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.text import RichText, Text
from openpyxl.chart.title import Title
from openpyxl.drawing.text import (
    CharacterProperties,
    ListStyle,
    Paragraph,
    ParagraphProperties,
    RegularTextRun,
    RichTextProperties,
)
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.benchmark import compare_to_industry
from src.load import load_clean_dataset
from src.ratios import compute_ratios

ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_DIR = ROOT / "reports" / "templates"
OUTPUT_DIR = ROOT / "reports" / "output"

# formatovanie pre excel
HEADER_FILL = PatternFill("solid", fgColor="4C72B0")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)

# nazvy ukazovatelov po slovensky
RATIO_LABELS = {
    "liq_cash":           "Likvidita 1. stupňa",
    "liq_quick":          "Likvidita 2. stupňa",
    "liq_current":        "Likvidita 3. stupňa",
    "dbt_total":          "Celková zadlženosť",
    "dbt_equity_ratio":   "Koef. samofinancovania",
    "dbt_debt_equity":    "Cudzie zdroje / vl. imanie",
    "prof_roa":           "ROA",
    "prof_roe":           "ROE",
    "prof_ros":           "ROS",
    "prof_ebitda_margin": "EBITDA marža",
    "act_asset_turnover": "Obrátka aktív",
    "altman_z_own":       "Altman Z (vlastný)",
    "altman_z_finstat":   "Altman Z (FinStat)",
}

# kategorie pre grafy v exceli
RATIO_CATEGORIES = [
    ("Likvidita",            ("liq_cash", "liq_quick", "liq_current")),
    ("Kapitálová štruktúra", ("dbt_total", "dbt_equity_ratio", "dbt_debt_equity")),
    ("Rentabilita",          ("prof_roa", "prof_roe", "prof_ros", "prof_ebitda_margin")),
    ("Aktivita a Altman Z",  ("act_asset_turnover", "altman_z_own", "altman_z_finstat")),
]


def _chart_title(text, size_hundredths_pt=1400, bold=True):
    """Nadpis grafu s explicitnou velkostou fontu."""
    cp = CharacterProperties(sz=size_hundredths_pt, b=bold)
    rich = RichText(p=[Paragraph(r=[RegularTextRun(rPr=cp, t=text)])])
    return Title(tx=Text(rich=rich), overlay=False)


def _font_props(size_hundredths_pt):
    """Nastavenie fontu pre osy a data labels v grafe."""
    cp = CharacterProperties(sz=size_hundredths_pt)
    pp = ParagraphProperties(defRPr=cp)
    return RichText(
        bodyPr=RichTextProperties(),
        lstStyle=ListStyle(),
        p=[Paragraph(pPr=pp, endParaRPr=cp)],
    )


def _meta_for(df, ico):
    """Ziska metadaje o firme podla ICO."""
    row = df.loc[df["Ico"] == ico]
    if row.empty:
        raise KeyError(f"Ico {ico!r} not found in dataset.")
    r = row.iloc[0]
    return {
        "ico":           ico,
        "nazov":         r["Nazov"],
        "dic":           r.get("DIC"),
        "odvetvie":      r["Odvetvie"],
        "kraj":          r["Kraj"],
        "okres":         r.get("Okres"),
        "mesto":         r.get("Mesto"),
        "pravna_forma":  r.get("Právna forma"),
        "kar_status":    r.get("KaR, Likvidácie"),
        "altman_finstat": r.get("Kreditný model: Altmanovo Z-skóre"),
        "altman_label":  r.get("Kreditný model: Altmanovo Z-skóre - indikácia"),
        "tržby":         r.get("Tržby (spolu)"),
        "ebitda":        r.get("EBITDA"),
        "zamestnanci":   r.get("Kategória zamestnancov(zo Štatistického úradu)"),
    }


def write_excel(meta, comparison, out_path):
    """Zapise benchmark do Excel suboru s grafmi."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- List 1: Suhrn ---
    ws = wb.active
    ws.title = "Súhrn"
    ws["A1"] = f"Benchmark report — {meta['nazov']}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    info_rows = [
        ("IČO",             meta["ico"]),
        ("DIČ",             meta["dic"]),
        ("Odvetvie",        meta["odvetvie"]),
        ("Kraj",            meta["kraj"]),
        ("Okres",           meta["okres"]),
        ("Mesto",           meta["mesto"]),
        ("Právna forma",    meta["pravna_forma"]),
        ("Veľkosť (zamestnanci)", meta["zamestnanci"]),
        ("Tržby (spolu)",   meta["tržby"]),
        ("EBITDA",          meta["ebitda"]),
        ("Altman Z (FinStat)", meta["altman_finstat"]),
        ("Indikácia FinStat", meta["altman_label"]),
        ("Stav (KaR / likvidácie)", meta["kar_status"]),
    ]
    for i, (label, value) in enumerate(info_rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 50

    # --- List 2: Porovnanie s odvetvim ---
    ws2 = wb.create_sheet("Porovnanie s odvetvím")
    headers = ["Ukazovateľ", "Firma", "Medián odvetvia", "Q25", "Q75", "n_peers", "Pozícia"]
    for col, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    code_to_row = {}
    for r_idx, row in enumerate(comparison.itertuples(index=False), start=2):
        code = row[0]
        ws2.cell(row=r_idx, column=1, value=RATIO_LABELS.get(code, code))
        for col, value in ((2, float(row[1])), (3, float(row[2])),
                           (4, float(row[3])), (5, float(row[4]))):
            cell = ws2.cell(row=r_idx, column=col, value=value)
            cell.number_format = "0.00"
        ws2.cell(row=r_idx, column=6, value=int(row[5]))
        ws2.cell(row=r_idx, column=7, value=row[6])
        code_to_row[code] = r_idx

    for col_idx, col_letter in enumerate("ABCDEFG", start=1):
        ws2.column_dimensions[col_letter].width = [26, 12, 16, 12, 12, 10, 26][col_idx - 1]

    # grafy - jeden per kategoria ukazovatelov
    tick_font = _font_props(1000)
    label_font = _font_props(1000)

    anchor_rows = [1, 26, 51, 76]
    for (cat_name, codes), anchor_row in zip(RATIO_CATEGORIES, anchor_rows):
        cat_rows = sorted(code_to_row[c] for c in codes if c in code_to_row)
        if not cat_rows:
            continue
        min_r, max_r = cat_rows[0], cat_rows[-1]
        if max_r - min_r + 1 != len(cat_rows):
            continue

        chart = BarChart()
        chart.type = "col"
        chart.style = 11
        chart.title = _chart_title(cat_name, size_hundredths_pt=1400, bold=True)
        chart.x_axis.txPr = tick_font
        chart.y_axis.txPr = tick_font
        chart.legend.position = "b"
        chart.dataLabels = DataLabelList(
            showVal=True,
            numFmt="0.00",
            dLblPos="outEnd",
            txPr=label_font,
        )

        firma_ref  = Reference(ws2, min_col=2, max_col=2, min_row=min_r, max_row=max_r)
        median_ref = Reference(ws2, min_col=3, max_col=3, min_row=min_r, max_row=max_r)
        cats_ref   = Reference(ws2, min_col=1, max_col=1, min_row=min_r, max_row=max_r)
        chart.add_data(firma_ref, titles_from_data=False)
        chart.add_data(median_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        chart.series[0].tx = SeriesLabel(v="Firma")
        chart.series[1].tx = SeriesLabel(v="Medián odvetvia")

        chart.height = 11
        chart.width = 22
        ws2.add_chart(chart, f"I{anchor_row}")

    wb.save(out_path)
    return out_path


def write_html(meta, comparison, out_path):
    """Vygeneruje HTML report z jinja2 sablony."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    env = Environment(
        loader=FileSystemLoader(TEMPLATE_DIR),
        autoescape=select_autoescape(["html"]),
    )
    template = env.get_template("company_report.html.j2")

    rows = comparison.to_dict(orient="records")
    n_peers_total = int(comparison["n_peers"].max()) if "n_peers" in comparison.columns else 0
    rendered = template.render(
        rows=rows,
        n_peers_total=n_peers_total,
        **meta,
    )
    out_path.write_text(rendered, encoding="utf-8")
    return out_path


def generate_report(ico, formats=("xlsx", "html"), output_dir=OUTPUT_DIR):
    """Hlavna funkcia - nacita data a vygeneruje reporty."""
    df = load_clean_dataset()
    ratios = compute_ratios(df)

    meta = _meta_for(df, ico)
    comparison = compare_to_industry(ratios, ico)

    paths = []
    safe_name = "".join(c if c.isalnum() else "_" for c in meta["nazov"])[:40]
    base = output_dir / f"benchmark_{ico}_{safe_name}"

    if "xlsx" in formats:
        paths.append(write_excel(meta, comparison, base.with_suffix(".xlsx")))
    if "html" in formats:
        paths.append(write_html(meta, comparison, base.with_suffix(".html")))
    return paths


def main():
    parser = argparse.ArgumentParser(description="Generovanie benchmark reportu pre firmu.")
    parser.add_argument("--ico", required=True, help="ICO firmy (8 cislic)")
    parser.add_argument("--output", type=Path, default=OUTPUT_DIR)
    parser.add_argument("--no-html", action="store_true", help="Preskocit HTML export")
    parser.add_argument("--no-xlsx", action="store_true", help="Preskocit Excel export")
    args = parser.parse_args()

    fmts = []
    if not args.no_xlsx:
        fmts.append("xlsx")
    if not args.no_html:
        fmts.append("html")

    paths = generate_report(args.ico, tuple(fmts), args.output)
    for p in paths:
        print(f"saved: {p}")


if __name__ == "__main__":
    main()
